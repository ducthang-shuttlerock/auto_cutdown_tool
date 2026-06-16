import os
import sys
import queue
import threading
import subprocess
import csv
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import ffmpeg

def parse_time(time_value):
    fps=25

    if time_value is None:
        return None

    time_str = str(time_value).strip()
    if not time_str:
        return None

    try:
        return float(time_str)
    except ValueError:
        pass

    parts = time_str.split(":")

    try:
        if len(parts) == 2:
            mm, ss = parts
            return int(mm) * 60 + float(ss)

        if len(parts) == 3:
            hh, mm, ss = parts
            return int(hh) * 3600 + int(mm) * 60 + float(ss)

        if len(parts) == 4:
            hh, mm, ss, ff = parts
            return int(hh) * 3600 + int(mm) * 60 + int(ss) + (int(ff) / fps)
    except ValueError:
        return None

    return None

def get_ffmpeg_path():
    if getattr(sys, "frozen", False):
        # PyInstaller bundles files in sys._MEIPASS (like the _internal folder)
        base_path = getattr(sys, "_MEIPASS", os.path.dirname(sys.executable))
        ffmpeg_path = os.path.join(base_path, "ffmpeg.exe")
        if os.path.exists(ffmpeg_path):
            return ffmpeg_path
        # Fallback to next to the executable
        return os.path.join(os.path.dirname(sys.executable), "ffmpeg.exe")
    else:
        base_path = os.path.dirname(__file__)
        return os.path.join(base_path, "ffmpeg.exe")


def open_folder(path):
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])
    except Exception:
        pass


class VideoCutterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Auto Cutdown Tool")
        self.root.geometry("760x560")
        self.root.minsize(720, 540)
        self.suffix_value = tk.StringVar(value="_cut")

        self.sheet_path = tk.StringVar()
        self.video_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.status_text = tk.StringVar(value="Ready.")
        self.progress_value = tk.DoubleVar(value=0)

        self.message_queue = queue.Queue()
        self.is_processing = False

        self.build_ui()
        self.poll_queue()

    def build_ui(self):
        main_frame = ttk.Frame(self.root, padding=16)
        main_frame.pack(fill="both", expand=True)

        title_label = ttk.Label(
            main_frame,
            text="Auto Cutdown Tool",
            font=("Segoe UI", 18, "bold")
        )
        title_label.pack(anchor="w", pady=(0, 12))

        description_label = ttk.Label(
            main_frame,
            text="Select an input CSV/Excel file, select a source video, choose an output folder, then click Start Processing.",
            wraplength=700
        )
        description_label.pack(anchor="w", pady=(0, 16))

        sheet_frame = ttk.LabelFrame(main_frame, text="Input Sheet", padding=12)
        sheet_frame.pack(fill="x", pady=(0, 10))

        sheet_entry = ttk.Entry(sheet_frame, textvariable=self.sheet_path)
        sheet_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))

        sheet_button = ttk.Button(
            sheet_frame,
            text="Browse...",
            command=self.choose_sheet_file
        )
        sheet_button.pack(side="left")

        video_frame = ttk.LabelFrame(main_frame, text="Source Video", padding=12)
        video_frame.pack(fill="x", pady=(0, 10))

        video_entry = ttk.Entry(video_frame, textvariable=self.video_path)
        video_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))

        video_button = ttk.Button(
            video_frame,
            text="Browse...",
            command=self.choose_video_file
        )
        video_button.pack(side="left")

        output_frame = ttk.LabelFrame(main_frame, text="Output Folder", padding=12)
        output_frame.pack(fill="x", pady=(0, 10))

        output_entry = ttk.Entry(output_frame, textvariable=self.output_path)
        output_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))

        output_button = ttk.Button(
            output_frame,
            text="Browse...",
            command=self.choose_output_folder
        )
        output_button.pack(side="left")

        options_frame = ttk.LabelFrame(main_frame, text="Options", padding=12)
        options_frame.pack(fill="x", pady=(0, 12))

        suffix_label = ttk.Label(options_frame, text="Output Suffix:")
        suffix_label.pack(side="left")

        suffix_entry = ttk.Entry(
            options_frame,
            textvariable=self.suffix_value,
            width=20
        )
        suffix_entry.pack(side="left", padx=(8, 0))

        actions_frame = ttk.Frame(main_frame)
        actions_frame.pack(fill="x", pady=(0, 10))

        self.start_button = ttk.Button(
            actions_frame,
            text="Start Processing",
            command=self.start_processing
        )
        self.start_button.pack(side="left")

        self.progress_bar = ttk.Progressbar(
            main_frame,
            variable=self.progress_value,
            maximum=100,
            mode="determinate"
        )
        self.progress_bar.pack(fill="x", pady=(8, 8))

        status_label = ttk.Label(main_frame, textvariable=self.status_text)
        status_label.pack(anchor="w", pady=(0, 10))

        log_frame = ttk.LabelFrame(main_frame, text="Log", padding=8)
        log_frame.pack(fill="both", expand=True)

        self.log_text = tk.Text(log_frame, wrap="word", state="disabled", height=16)
        self.log_text.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        scrollbar.pack(side="right", fill="y")
        self.log_text.configure(yscrollcommand=scrollbar.set)

    def choose_sheet_file(self):
        path = filedialog.askopenfilename(
            title="Select CSV or Excel file",
            filetypes=[
                ("CSV and Excel files", "*.csv *.xlsx *.xls"),
                ("CSV files", "*.csv"),
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*"),
            ]
        )
        if path:
            self.sheet_path.set(path)
            self.log(f"Selected input sheet: {path}")

    def choose_video_file(self):
        path = filedialog.askopenfilename(
            title="Select source video",
            filetypes=[
                ("Video files", "*.mp4 *.mov *.mxf *.avi *.mkv *.wmv"),
                ("All files", "*.*"),
            ]
        )
        if path:
            self.video_path.set(path)
            self.log(f"Selected source video: {path}")

    def choose_output_folder(self):
        path = filedialog.askdirectory(title="Select output folder")
        if path:
            self.output_path.set(path)
            self.log(f"Selected output folder: {path}")

    def set_processing_state(self, processing):
        self.is_processing = processing
        self.start_button.configure(state="disabled" if processing else "normal")

    def start_processing(self):
        if self.is_processing:
            return

        sheet_file = self.sheet_path.get().strip()
        video_file = self.video_path.get().strip()
        output_dir = self.output_path.get().strip()
        suffix = self.suffix_value.get().strip()

        if not suffix:
            suffix = "_cut"

        if not sheet_file:
            messagebox.showwarning("Missing Information", "Please select an input sheet.")
            return

        if not video_file:
            messagebox.showwarning("Missing Information", "Please select a source video.")
            return

        if not output_dir:
            messagebox.showwarning("Missing Information", "Please select an output folder.")
            return

        if not os.path.exists(sheet_file):
            messagebox.showerror("Error", "Input sheet not found.")
            return

        if not os.path.exists(video_file):
            messagebox.showerror("Error", "Source video not found.")
            return

        self.progress_value.set(0)
        self.status_text.set("Processing...")
        self.log("=" * 60)
        self.log("Starting processing...")

        self.set_processing_state(True)

        worker = threading.Thread(
            target=self.process_videos,
            args=(sheet_file, video_file, output_dir, suffix),
            daemon=True
        )
        worker.start()

    def process_videos(self, sheet_file, video_file, output_dir, suffix):
        try:
            if sheet_file.lower().endswith(".csv"):
                with open(sheet_file, newline='', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    rows = list(reader)
                    fieldnames = reader.fieldnames

            elif sheet_file.lower().endswith(".xlsx"):
                wb = load_workbook(sheet_file, data_only=True)
                ws = wb.active

                data = list(ws.iter_rows(values_only=True))

                if not data:
                    self.message_queue.put(("error", "Excel file is empty."))
                    return

                fieldnames = [str(cell).strip() if cell else "" for cell in data[0]]

                rows = []
                for r in data[1:]:
                    row_dict = {}
                    for i, col in enumerate(fieldnames):
                        row_dict[col] = r[i] if i < len(r) else None
                    rows.append(row_dict)

            else:
                self.message_queue.put(("error", "Only CSV and XLSX files are supported."))
                return

            if not rows:
                self.message_queue.put(("error", "Input file is empty."))
                return

            columns_lower = {str(col).strip().lower(): col for col in fieldnames}

            supported_formats = [
                ("source_in", "source_out"),
                ("source in", "source out"),
                ("timeline in", "timeline out"),
            ]

            selected_in_col = None
            selected_out_col = None

            for in_name, out_name in supported_formats:
                if in_name in columns_lower and out_name in columns_lower:
                    selected_in_col = columns_lower[in_name]
                    selected_out_col = columns_lower[out_name]
                    break

            if not selected_in_col or not selected_out_col:
                self.message_queue.put((
                    "error",
                    "Missing required columns. Supported formats:\n"
                    "- source_in, source_out\n"
                    "- Source In, Source Out\n"
                    "- Timeline In, Timeline Out"
                ))
                return

            if not rows:
                self.message_queue.put(("error", "The input sheet is empty."))
                return

            os.makedirs(output_dir, exist_ok=True)

            total_rows = len(rows)
            success_count = 0
            error_count = 0
            error_messages = []

            video_base_name, video_ext = os.path.splitext(os.path.basename(video_file))

            for index, row in enumerate(rows):
                row_number = index + 2
                source_in_raw = row.get(selected_in_col)
                source_out_raw = row.get(selected_out_col)

                self.message_queue.put((
                    "log",
                    f"[Row {row_number}] Processing clip..."
                ))

                try:
                    source_in = parse_time(source_in_raw)
                    source_out = parse_time(source_out_raw)

                    if source_in is None or source_out is None:
                        raise ValueError("Invalid source_in/source_out value.")

                    if source_in >= source_out:
                        raise ValueError("source_in must be smaller than source_out.")

                    safe_suffix = suffix if suffix.startswith("_") else "_" + suffix
                    out_filename = f"{video_base_name}{safe_suffix}{index+1}.mp4"
                    out_filepath = os.path.join(output_dir, out_filename)

                    # with VideoFileClip(video_file) as video:
                    #     duration = float(video.duration)

                    #     if source_in > duration:
                    #         raise ValueError(
                    #             f"Start time ({source_in:.2f}s) exceeds video duration ({duration:.2f}s)."
                    #         )

                    #     if source_out > duration:
                    #         self.message_queue.put((
                    #             "log",
                    #             f"[Row {row_number}] End time exceeds video duration. Clamped to {duration:.2f}s."
                    #         ))
                    #         source_out = duration

                    #     clip = video.subclipped(source_in, source_out)
                    #     clip.write_videofile(
                    #         out_filepath,
                    #         codec="libx264",
                    #         audio_codec="aac",
                    #         logger=None
                    #     )
                    #     clip.close()                                         
                    ffmpeg_path = get_ffmpeg_path()
                    self.message_queue.put(("log", f"[Debug] ffmpeg path: {ffmpeg_path}"))
                    self.message_queue.put(("log", f"[Debug] ffmpeg exists: {os.path.exists(ffmpeg_path)}"))

                                            
                    start = source_in
                    duration = source_out - source_in

                    command = [
                        ffmpeg_path,
                        "-y",
                        "-ss", str(start),
                        "-i", video_file,
                        "-t", str(duration),
                        "-c", "copy",
                        out_filepath
                    ]

                    result = subprocess.run(
                        command,
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.PIPE,
                        text=True
                    )

                    if result.returncode != 0:
                        # Fallback to transcoding (re-encoding) if stream copy fails
                        # (e.g. incompatible codecs like ProRes to MP4 container)
                        self.message_queue.put((
                            "log",
                            f"[Row {row_number}] Fast copy failed (incompatible codecs). Falling back to transcoding..."
                        ))
                        fallback_command = [
                            ffmpeg_path,
                            "-y",
                            "-ss", str(start),
                            "-i", video_file,
                            "-t", str(duration),
                            "-c:v", "libx264",
                            "-c:a", "aac",
                            "-pix_fmt", "yuv420p",
                            out_filepath
                        ]
                        result = subprocess.run(
                            fallback_command,
                            stdout=subprocess.DEVNULL,
                            stderr=subprocess.PIPE,
                            text=True
                        )
                        if result.returncode != 0:
                            raise RuntimeError(result.stderr.strip() or "FFmpeg transcoding failed.")
                
                    success_count += 1
                    self.message_queue.put((
                        "log",
                        f"[Row {row_number}] Success -> {out_filepath}"
                    ))

                except Exception as e:
                    if os.path.exists(out_filepath):
                        try:
                            os.remove(out_filepath)
                        except Exception:
                            pass
                    error_count += 1
                    error_message = f"[Row {row_number}] Error: {e}"
                    error_messages.append(error_message)
                    self.message_queue.put(("log", error_message))

                progress = ((index + 1) / total_rows) * 100
                self.message_queue.put(("progress", progress))
                self.message_queue.put((
                    "status",
                    f"Processed {index + 1}/{total_rows} rows..."
                ))

            summary_lines = [
                "Processing completed.",
                f"Success: {success_count}",
                f"Errors: {error_count}",
                f"Output folder: {output_dir}"
            ]

            if error_messages:
                summary_lines.append("")
                summary_lines.append("Some errors:")
                summary_lines.extend(error_messages[:10])

            self.message_queue.put(("done", "\n".join(summary_lines)))
            self.message_queue.put(("open_folder", output_dir))

        except Exception as e:
            self.message_queue.put(("error", f"Unexpected error: {e}"))

    def poll_queue(self):
        try:
            while True:
                message_type, payload = self.message_queue.get_nowait()

                if message_type == "log":
                    self.log(payload)
                elif message_type == "progress":
                    self.progress_value.set(payload)
                elif message_type == "status":
                    self.status_text.set(payload)
                elif message_type == "done":
                    self.set_processing_state(False)
                    self.status_text.set("Completed.")
                    self.log("-" * 60)
                    self.log(payload)
                    messagebox.showinfo("Result", payload)
                elif message_type == "error":
                    self.set_processing_state(False)
                    self.status_text.set("Error.")
                    self.log(f"ERROR: {payload}")
                    messagebox.showerror("Error", payload)
                elif message_type == "open_folder":
                    open_folder(payload)
        except queue.Empty:
            pass

        self.root.after(100, self.poll_queue)

    def log(self, message):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", message + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")


def main():
    root = tk.Tk()
    app = VideoCutterApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()